from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.template.defaultfilters import timesince
from django.utils.timezone import now 
from attendance.models import Attendance
from leave.models import Holiday as LeaveHoliday, Leave, LeaveBalance
from resignation.models import Resignation 
from .models import Admin, AllowedDomain, Employee ,EmployeeDocument, Location, Department, Designation, MessageCategory, MessageSubType, Role ,ProbationConfiguration,EmployeeWarning, YsMenuLinkMaster, YsMenuMaster, YsMenuRoleMaster,CelebrationWish
from .forms import AdminForm, AllowedDomainForm, LocationForm, DepartmentForm, DesignationForm, RoleForm,EmployeeWarningForm
from datetime import date, datetime, time, timedelta
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q ,Count,Sum
from .utils import authenticate_user, get_domain_restriction_message, get_user_display_name, simple_hash, set_employee_password, validate_email_domain
import json
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from itertools import chain
from operator import attrgetter
from django.utils.timezone import localtime
import calendar
# Authentication decorator

def dynamic_menu(request):
    """Provide active menus and submenus for sidebar."""
    if not request.session.get('user_authenticated'):
        return {}

    menus = YsMenuMaster.objects.filter(status=True).order_by('seq')
    menu_data = []

    for menu in menus:
        submenus = YsMenuLinkMaster.objects.filter(menu=menu, status=True).order_by('seq')
        menu_data.append({
            'id': menu.menu_id,
            'name': menu.menu_name,
            'icon': menu.menu_icon,
            'url': menu.menu_url,
            'submenus': submenus
        })

    return {'menu_data': menu_data}

def login_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_authenticated'):
            messages.error(request, 'Please login to access this page.')
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

def role_required(allowed_roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            user_role = request.session.get('user_role')
            if not user_role or user_role not in allowed_roles:
                messages.error(request, 'You do not have permission to access this page.')
                return redirect('access_denied')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# Authentication Views
def login_view(request):
    if request.session.get('user_authenticated'):
        return redirect('dashboard')
    # Get domain restriction message for display
    domain_message = get_domain_restriction_message()
    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')
        # Validate email format
        if not email or '@' not in email:
            messages.error(request, 'Please enter a valid email address')
            return render(request, 'hr/login.html', {'domain_message': domain_message})
        
        # Validate email domain before authentication
        is_domain_allowed, domain_validation_message = validate_email_domain(email)
        
        if not is_domain_allowed:
            messages.error(request, f'Access denied: {domain_validation_message}')
            # Log the attempted login for security
            print(f"Domain blocked login attempt: {email} - Reason: {domain_validation_message}")
            return render(request, 'hr/login.html', {'domain_message': domain_message})
        user, user_type = authenticate_user(email, password)
        
        if user and user_type:
            request.session['user_authenticated'] = True
            request.session['user_email'] = email
            request.session['user_role'] = user_type

            # Debug print
            print(f"DEBUG: User Type: {user_type}")

            if user_type == 'SUPER ADMIN':
                request.session['user_department'] = "NONE"
                profile_picture = user.profile_picture
            else:
                request.session['user_department'] = user.department if user.department else None
                profile_picture = user.profile_picture.url if user.profile_picture else None

            request.session['user_id'] = getattr(user, 'admin_id', getattr(user, 'id', None))
            request.session['user_name'] = get_user_display_name(user, user_type)
            
            request.session['profile_picture'] = profile_picture  # Store in session
            
            # ✅ Store employee_id safely in session (used for approvals)
            if hasattr(user, 'employee_id'):
                request.session['employee_id'] = user.employee_id
            else:
                request.session['employee_id'] = None

            messages.success(request, f'Welcome back, {request.session["user_name"]}!')

            # Redirect based on role
            if user_type in ['ADMIN', 'HR', 'SUPER ADMIN','Branch Manager']:
                return redirect('dashboard')
            
            else:
                return redirect('employee_dashboard')

        else:
            messages.error(request, 'Invalid email or password.')
    
    # Default (GET request) - pass domain message to template
    return render(request, 'hr/login.html', {'domain_message': domain_message})

def logout_view(request):
    request.session.flush()
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')

@login_required
def change_password(request):
    user_email = request.session.get('user_email')
    user_role = request.session.get('user_role')
    
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Verify current password
        user, user_type = authenticate_user(user_email, current_password)
        if not user:
            messages.error(request, 'Current password is incorrect.')
            return render(request, 'hr/change_password.html')
        
        # Check new password requirements
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
        elif len(new_password) < 6:
            messages.error(request, 'Password must be at least 6 characters long.')
        else:
            try:
                # Update password based on user type
                if user_role == 'SUPER ADMIN':
                    admin = Admin.objects.get(email=user_email)
                    admin.password_hash = simple_hash(new_password)
                    admin.updated_at = timezone.now()
                    admin.save()
                    messages.success(request, 'Password changed successfully!')
                
                elif user_role in ['EMPLOYEE', 'MANAGER', 'HR', 'ADMIN']:
                    # For employees and other roles
                    employee = Employee.objects.get(email=user_email, status='active')
                    set_employee_password(employee, new_password)
                    messages.success(request, 'Password changed successfully!')
                
                # Redirect based on user role after successful password change
                # if user_role == 'EMPLOYEE':
                #     return redirect('employee_dashboard')
                # elif user_role in ['ADMIN', 'HR', 'SUPER ADMIN', 'MANAGER']:
                #     return redirect('dashboard')
                if user_role == 'SUPER ADMIN':
                    return redirect('dashboard')
                else:
                    return redirect('employee_dashboard')
                    
            except Employee.DoesNotExist:
                messages.error(request, 'Employee account not found or inactive.')
            except Exception as e:
                messages.error(request, f'Error changing password: {str(e)}')
        
        # If we get here, there was an error - stay on the same page
        return render(request, 'hr/change_password.html')
    
    # GET request - show the form
    context = {
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
        'today_date': date.today(),
    }
    return render(request, 'hr/change_password.html', context)

def access_denied(request):
    return render(request, 'hr/access_denied.html')

# Dashboard Views
@login_required
def dashboard(request):
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')
    
    # Get branch manager location if applicable
    current_branch_manager_location = None
    if user_role == 'BRANCH MANAGER':
        try:
            current_branch_manager = Employee.objects.get(email=user_email)
            current_branch_manager_location = current_branch_manager.location
        except Employee.DoesNotExist:
            pass

    # ---- Base Data ----
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        total_employees = Employee.objects.count()
        total_location = Location.objects.count()
        # Location-wise count
        location_data = Employee.objects.values('location').annotate(count=Count('id'))
        # Department-wise count
        department_data = Employee.objects.values('department').annotate(count=Count('id'))
    elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
        total_employees = Employee.objects.filter(location__iexact=current_branch_manager_location).count()
        total_location = 1  # Branch manager only sees their location
        # Location-wise count (only their location)
        location_data = Employee.objects.filter(location__iexact=current_branch_manager_location).values('location').annotate(count=Count('id'))
        # Department-wise count (only their location)
        department_data = Employee.objects.filter(location__iexact=current_branch_manager_location).values('department').annotate(count=Count('id'))
    else:
        total_employees = 0
        total_location = 0
        location_data = []
        department_data = []

    location_labels = [loc['location'] for loc in location_data if loc['location']]
    location_counts = [loc['count'] for loc in location_data if loc['location']]
    department_labels = [dept['department'] for dept in department_data if dept['department']]
    department_counts = [dept['count'] for dept in department_data if dept['department']]

    # ---- NEW: Attendance Data ----
    today = date.today()
    try:
        # Filter attendance records for today only
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            today_attendance = Attendance.objects.filter(date=today)
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            today_attendance = Attendance.objects.filter(
                date=today,
                employee__location__iexact=current_branch_manager_location
            )
        else:
            today_attendance = Attendance.objects.none()
        
        # Count employees who have checked in today (check_in is not null)
        today_present = today_attendance.filter(check_in__isnull=False).count()
        today_present_total = total_employees
        
        print(f"Today: {today}")
        print(f"Today attendance records: {today_attendance.count()}")
        print(f"Today present: {today_present}")
        
    except Exception as e:
        print(f"Attendance error: {e}")
        today_present = 0
        today_present_total = total_employees

    # ---- NEW: Resignation Data ----
    try:
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            total_resignations = Resignation.objects.filter(status__in=['applied', 'accepted']).count()
            pending_resignations = Resignation.objects.filter(status='applied').count()
            # Active notice period
            active_notice = Resignation.objects.filter(
                status='accepted',
                last_working_date__gte=today
            ).count()
            # Completed this month
            current_month = today.month
            current_year = today.year
            completed_this_month = Resignation.objects.filter(
                status='completed',
                last_working_date__month=current_month,
                last_working_date__year=current_year
            ).count()
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            total_resignations = Resignation.objects.filter(
                employee__location__iexact=current_branch_manager_location
            ).count()
            pending_resignations = Resignation.objects.filter(
                status='applied',
                employee__location__iexact=current_branch_manager_location
            ).count()
            active_notice = Resignation.objects.filter(
                status='accepted',
                last_working_date__gte=today,
                employee__location__iexact=current_branch_manager_location
            ).count()
            current_month = today.month
            current_year = today.year
            completed_this_month = Resignation.objects.filter(
                status='completed',
                last_working_date__month=current_month,
                last_working_date__year=current_year,
                employee__location__iexact=current_branch_manager_location
            ).count()
        else:
            total_resignations = 0
            pending_resignations = 0
            active_notice = 0
            completed_this_month = 0
        
    except Exception as e:
        print(f"Resignation error: {e}")
        total_resignations = 0
        pending_resignations = 0
        active_notice = 0
        completed_this_month = 0

    # ---- NEW: Leave Data ----
    try:
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            recent_leaves = Leave.objects.select_related('employee', 'leave_type').filter(
                start_date__lte=today,
                end_date__gte=today,
                status__in=['approved']
            ).order_by('-applied_date')[:5]
            
            pending_leave_requests = Leave.objects.filter(status__in=['pending', 'new']).count()
            approved_leaves_today = Leave.objects.filter(
                start_date__lte=today,
                end_date__gte=today,
                status='approved'
            ).count()
            
            total_leaves = Leave.objects.filter(
                start_date__lte=today,
                end_date__gte=today,
                status='approved').count()
                
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            recent_leaves = Leave.objects.select_related('employee', 'leave_type').filter(
                start_date__lte=today,
                end_date__gte=today,
                status__in=['approved'],
                employee__location__iexact=current_branch_manager_location
            ).order_by('-applied_date')[:5]
            
            pending_leave_requests = Leave.objects.filter(
                status__in=['pending', 'new'],
                employee__location__iexact=current_branch_manager_location
            ).count()
            
            approved_leaves_today = Leave.objects.filter(
                start_date__lte=today,
                end_date__gte=today,
                status='approved',
                employee__location__iexact=current_branch_manager_location
            ).count()
            
            total_leaves = Leave.objects.filter(
                start_date__lte=today,
                end_date__gte=today,
                status='approved',
                employee__location__iexact=current_branch_manager_location
            ).count()
        else:
            recent_leaves = []
            pending_leave_requests = 0
            approved_leaves_today = 0
            total_leaves = 0
        
    except Exception as e:
        print(f"Leave data error: {e}")
        recent_leaves = []
        pending_leave_requests = 0
        approved_leaves_today = 0
        total_leaves = 0

    # ---- NEW: Recent Resignations ----
    try:
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            recent_resignations = Resignation.objects.select_related('employee').filter(
                status__in=['applied', 'accepted']
            ).order_by('-resignation_date')[:3]
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            recent_resignations = Resignation.objects.select_related('employee').filter(
                employee__location__iexact=current_branch_manager_location,
                status__in=['applied', 'accepted']
            ).order_by('-resignation_date')[:3]
        else:
            recent_resignations = []
    except Exception as e:
        print(f"Recent resignations error: {e}")
        recent_resignations = []
    

    # ---- NEW: Location-wise Attendance Data ----
    location_present_counts = []
    location_absent_counts = []
    location_attendance_labels = []
    
    try:
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            # Get all locations that have employees
            locations_with_employees = Employee.objects.filter(
                location__isnull=False
            ).values_list('location', flat=True).distinct()
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            # Only show the branch manager's location
            locations_with_employees = [current_branch_manager_location]
        else:
            locations_with_employees = []
        
        print(f"Locations with employees: {list(locations_with_employees)}")
        
        for location in locations_with_employees:
            if location:  # Skip empty locations
                # Count total ACTIVE employees in this location
                if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
                    total_in_location = Employee.objects.filter(
                        location=location, 
                        status='active'
                    ).count()
                    
                    # Count present employees in this location TODAY
                    present_in_location = Attendance.objects.filter(
                        date=today,
                        check_in__isnull=False,
                        employee__location=location,
                        employee__status='active'
                    ).count()
                elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
                    total_in_location = Employee.objects.filter(
                        location__iexact=current_branch_manager_location, 
                        status='active'
                    ).count()
                    
                    present_in_location = Attendance.objects.filter(
                        date=today,
                        check_in__isnull=False,
                        employee__location__iexact=current_branch_manager_location,
                        employee__status='active'
                    ).count()
                else:
                    total_in_location = 0
                    present_in_location = 0
                
                # Calculate absent count for TODAY
                absent_in_location = total_in_location - present_in_location
                
                location_attendance_labels.append(location)
                location_present_counts.append(present_in_location)
                location_absent_counts.append(absent_in_location)
                
                print(f"Location: {location}, Total: {total_in_location}, Present: {present_in_location}, Absent: {absent_in_location}")
            
    except Exception as e:
        print(f"Location attendance error: {e}")
        # If no data found for today, show zeros
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            locations_with_employees = Employee.objects.filter(
                location__isnull=False
            ).values_list('location', flat=True).distinct()
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            locations_with_employees = [current_branch_manager_location]
        else:
            locations_with_employees = []
        
        for location in locations_with_employees:
            if location:
                if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
                    total_in_location = Employee.objects.filter(location=location, status='active').count()
                elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
                    total_in_location = Employee.objects.filter(location__iexact=current_branch_manager_location, status='active').count()
                else:
                    total_in_location = 0
                    
                location_attendance_labels.append(location)
                location_present_counts.append(0)  # No attendance data for today
                location_absent_counts.append(total_in_location)  # All absent today

    # ---- Shared Context ----
    context = {
        'total_employees': total_employees,
        'location_labels': json.dumps(location_labels),
        'location_counts': json.dumps(location_counts),
        'department_labels': json.dumps(department_labels),
        'department_counts': json.dumps(department_counts),
        'today_date': today.strftime("%d %B %Y"),
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
        'total_location': total_location,
        
        # NEW: Attendance Data
        'today_present': today_present,
        'today_present_total': today_present_total,
        
        # NEW: Resignation Data
        'total_resignations': total_resignations,
        'pending_resignations': pending_resignations,
        'active_notice': active_notice,
        'completed_this_month': completed_this_month,
        
        # NEW: Leave Data
        'recent_leaves': recent_leaves,
        'pending_leave_requests': pending_leave_requests,
        'approved_leaves_today': approved_leaves_today,
        'total_leaves': total_leaves,  # NEW: Total leave count
        
        # NEW: Recent Resignations
        'recent_resignations': recent_resignations,

        # NEW: Location-wise Attendance Data for Charts
        'location_attendance_labels': json.dumps(location_attendance_labels),
        'location_present_counts': json.dumps(location_present_counts),
        'location_absent_counts': json.dumps(location_absent_counts)
    }

    # ---- Role-Based Additions ----
    if user_role == 'ADMIN':
        total_admins = Admin.objects.count()
        new_admins = Admin.objects.order_by('-created_at')[:5]
        active_employees = Employee.objects.filter(status='active').count()

        context.update({
            'total_admins': total_admins,
            'new_admins': new_admins,
            'active_employees': active_employees,
        })
    else:
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
            active_employees = Employee.objects.filter(status='active').count()
        elif user_role == 'BRANCH MANAGER' and current_branch_manager_location:
            active_employees = Employee.objects.filter(
                status='active', 
                location__iexact=current_branch_manager_location
            ).count()
        else:
            active_employees = 0
            
        context.update({
            'active_employees': active_employees,
        })

    return render(request, 'hr/dashboard.html', context)

@login_required
def employee_dashboard(request):
    user_email = request.session.get('user_email')
    user_role = request.session.get('user_role')

    try:
        employee_profile = Employee.objects.get(email=user_email)
    except Employee.DoesNotExist:
        messages.warning(request, 'Employee profile not found.')
        return redirect('access_denied')

    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(employee=employee_profile, date=today).first()
    office_start_time = time(9, 30)  # 9:30 AM
    punctuality_status = None

    # ✅ If attendance exists, determine punctuality
    if today_attendance and today_attendance.check_in:
        check_in_time = timezone.localtime(today_attendance.check_in).time()
        punctuality_status = "On Time" if check_in_time <= office_start_time else "Late"
        
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'check_in':
            if today_attendance:
                messages.warning(request, 'You have already checked in today.')
            else:
                current_time = timezone.localtime().time()
                punctuality = "On Time" if current_time <= office_start_time else "Late"
                checkin_lat = request.POST.get('latitude')
                checkin_long = request.POST.get('longitude')
                checkin_address = request.POST.get('address')
                Attendance.objects.create(
                    employee=employee_profile,
                    date=today,
                    check_in=timezone.now(),
                    checkin_latitude=checkin_lat,
                    checkin_longitude=checkin_long,
                    checkin_address=checkin_address
                )
                messages.success(request, 'Check-in successful!')
                return redirect('employee_dashboard')

        elif action == 'check_out':
            if not today_attendance:
                messages.error(request, 'You need to check in first.')
            elif today_attendance.check_out:
                messages.warning(request, 'You have already checked out today.')
            else:
                 # ✅ Get checkout location data
                checkout_lat = request.POST.get('latitude')
                checkout_long = request.POST.get('longitude')
                checkout_address = request.POST.get('address')

                today_attendance.check_out = timezone.now()
                today_attendance.checkout_latitude = checkout_lat
                today_attendance.checkout_longitude = checkout_long
                today_attendance.checkout_address = checkout_address
                today_attendance.save()

                messages.success(request, 'Check-out successful!')
                return redirect('employee_dashboard')

    total_team_members = None
    if employee_profile.department:
        current_manager = Employee.objects.get(email=user_email)
        if user_role in ['MANAGER','TL']:
            total_team_members = Employee.objects.filter(
                    Q(reporting_manager_id=current_manager.id) |
                    Q(reporting_manager__icontains=current_manager.first_name)
                    ).order_by('first_name').count()
        else:
            total_team_members = Employee.objects.filter(
            department__iexact=employee_profile.department,
            status__iexact='active').count()
    current_year = today.year
    total_remaining_leaves = LeaveBalance.objects.filter(
        employee=employee_profile,
        year=current_year
    ).aggregate(
        total=Sum('leaves_remaining')
    )['total'] or 0
    
    next_30_days = today + timedelta(days=30)
    upcoming_holidays = LeaveHoliday.objects.filter(
        date__range=(today, next_30_days),
        region__name=employee_profile.location 
    ).count() or 0
    
    # === Productivity metrics for charts ===
    def calculate_hours(attendance_obj):
        """Return worked hours for a single attendance record."""
        if not attendance_obj or not attendance_obj.check_in:
            return 0

        end_time = attendance_obj.check_out or timezone.now()
        duration = (timezone.localtime(end_time) - timezone.localtime(attendance_obj.check_in)).total_seconds()
        return round(max(duration, 0) / 3600, 2)

    # Last 7 days (including today)
    weekly_labels = []
    weekly_hours = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        attendance_obj = Attendance.objects.filter(employee=employee_profile, date=day).first()
        weekly_labels.append(day.strftime('%a'))
        weekly_hours.append(calculate_hours(attendance_obj))

    worked_days = [h for h in weekly_hours if h > 0]
    average_work_hours = round(sum(worked_days) / len(worked_days), 2) if worked_days else 0
    lost_days = weekly_hours.count(0)

    # Last 6 months totals
    monthly_labels = []
    monthly_hours = []
    current_month = today.month
    current_year = today.year

    for i in range(5, -1, -1):
        month = current_month - i
        year = current_year
        # adjust year/month when subtracting across year boundary
        while month <= 0:
            month += 12
            year -= 1

        start_date = date(year, month, 1)
        _, last_day = calendar.monthrange(year, month)
        end_date = date(year, month, last_day)

        monthly_labels.append(start_date.strftime('%b'))

        month_attendance = Attendance.objects.filter(
            employee=employee_profile,
            date__range=(start_date, end_date),
            check_in__isnull=False
        )

        total_seconds = 0
        for att in month_attendance:
            if not att.check_out:
                continue
            total_seconds += (timezone.localtime(att.check_out) - timezone.localtime(att.check_in)).total_seconds()

        monthly_hours.append(round(max(total_seconds, 0) / 3600, 1))

    # ✅ Get only recent warnings & appreciations (last 7 days)
    seven_days_ago = today - timedelta(days=7)

    recent_messages = EmployeeWarning.objects.filter(
        employee_code=employee_profile.employee_id,
        warning_date__gte=seven_days_ago
    ).order_by('-warning_date')

    # Same messages used for dashboard + popup
    notifications = recent_messages
    notifications_count = recent_messages.count()

    context = {
        'employee': employee_profile,
        'today_date': today,
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
        'total_team_members': total_team_members,
        'upcoming_holidays': upcoming_holidays,
        'total_remaining_leaves': total_remaining_leaves,
        'today_attendance': today_attendance,
        'punctuality_status': punctuality_status,
        'weekly_labels': weekly_labels,
        'weekly_hours': weekly_hours,
        'average_work_hours': average_work_hours,
        'lost_days': lost_days,
        'monthly_labels': monthly_labels,
        'monthly_hours': monthly_hours,
        'recent_messages': recent_messages,
        'recent_messages_count': recent_messages.count(),

        # Notification bell popup = same messages
        'notifications': notifications,
        'notifications_count': notifications_count,
    }
    return render(request, 'hr/employee_dashboard.html', context)

@login_required
def total_team_members(request):
    user_email = request.session.get('user_email')
    user_role = request.session.get('user_role')
    current_manager = Employee.objects.get(email=user_email)
   
    try:
        employee = Employee.objects.get(email=user_email)
    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect('access_denied')
    if user_role in ['MANAGER','TL']:
        
        # ✅ Fetch all employees in the same department
        team_members = Employee.objects.filter(
                    Q(reporting_manager_id=current_manager.id) |
                    Q(reporting_manager__icontains=current_manager.first_name)
                    ).order_by('first_name').exclude(id=employee.id)  # exclude self
    else:
        team_members = Employee.objects.filter(department__iexact=employee.department).exclude(id=employee.id)  # exclude self
    context = {
        'employee': employee,
        'team_members': team_members,
    }
    return render(request, 'hr/totalteam-member.html',context)
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def add_employee(request):
    # Get all active managers for the dropdown
    managers = Employee.objects.filter(
        role__in=['Manager', 'TL', 'HR', 'Admin', 'Super Admin'],
        status='active'
    ).order_by('first_name', 'last_name')
   
    # Get all active locations for the dropdown
    locations = Location.objects.filter(is_active=True).order_by('name')
    
    # Get all active departments for the dropdown
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Get all active designations for the dropdown - use 'title' instead of 'name'
    designations = Designation.objects.filter(is_active=True).order_by('title')

    # Get all active designations for the dropdown - use 'title' instead of 'name'
    roles = Role.objects.filter(is_active=True).order_by('name')

    if request.method == "POST":
        try:
            # Check if employee ID already exists
            employee_id = request.POST.get('employee_id')
            if Employee.objects.filter(employee_id=employee_id).exists():
                messages.error(request, f"Employee with ID {employee_id} already exists.")
                return render(request, 'hr/add_employee.html', {
                    'managers': managers,
                    'locations': locations,
                    'departments': departments,
                    'designations': designations,
                    'roles':roles,
                })
           
            # Check if email already exists
            email = request.POST.get('email')
            if Employee.objects.filter(email=email).exists():
                messages.error(request, f"Employee with email {email} already exists.")
                return render(request, 'hr/add_employee.html', {
                    'managers': managers,
                    'locations': locations,
                    'departments': departments,
                    'designations': designations,
                    'roles':roles,
                })
           
            # Extract reporting manager ID from the selected option
            reporting_manager_full = request.POST.get('reporting_manager')
            reporting_manager_id = None
           
            if reporting_manager_full:
                try:
                    name_part = reporting_manager_full.split(' (')[0]
                    first_name, last_name = name_part.split(' ', 1) if ' ' in name_part else (name_part, '')
                   
                    manager = Employee.objects.filter(
                        first_name=first_name,
                        last_name=last_name,
                        status='active'
                    ).first()
                   
                    if manager:
                        reporting_manager_id = manager.employee_id
                except (IndexError, ValueError):
                    reporting_manager_id = None
            
            # Handle date fields - FIXED: Use correct format for DD-MM-YYYY
            def parse_date(date_str, default=None):
                if date_str:
                    try:
                        # Try parsing as DD-MM-YYYY
                        return datetime.strptime(date_str, "%d-%m-%Y").date()
                    except ValueError:
                        try:
                            # Fallback: try YYYY-MM-DD (in case browser auto-formatting)
                            return datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            return default
                return default

            # Parse all date fields
            date_of_joining = parse_date(request.POST.get('date_of_joining'), timezone.now().date())
            date_of_birth = parse_date(request.POST.get('date_of_birth'))
            contract_end_date = parse_date(request.POST.get('contract_end_date'))
            marriage_date = parse_date(request.POST.get('marriage_date'))
            pf_joining_date = parse_date(request.POST.get('pf_joining_date'))
            policy_start_date = parse_date(request.POST.get('policy_start_date'))
            policy_end_date = parse_date(request.POST.get('policy_end_date'))

            # Handle location
            location_id = request.POST.get('location')
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                    location = location_instance.name
                except Location.DoesNotExist:
                    location_instance = None
                    location = None

            # Handle department
            department_id = request.POST.get('department')
            department_instance = None
            if department_id:
                try:
                    department_instance = Department.objects.get(id=department_id)
                    department = department_instance.name
                except Department.DoesNotExist:
                    department_instance = None
                    department = None

            # Handle designation
            designation_id = request.POST.get('designation')
            designation_instance = None
            if designation_id:
                try:
                    designation_instance = Designation.objects.get(id=designation_id)
                    designation = designation_instance.title
                except Designation.DoesNotExist:
                    designation_instance = None
                    designation = None

            # Handle basic salary
            basic_salary_str = request.POST.get('basic_salary')
            basic_salary = None
            if basic_salary_str:
                try:
                    basic_salary = float(basic_salary_str)
                except (ValueError, TypeError):
                    basic_salary = 0.00

            # Handle boolean fields
            physically_handicapped = request.POST.get('physically_handicapped') == 'on'

            # Handle coverage amount
            coverage_amount_str = request.POST.get('coverage_amount')
            coverage_amount = None
            if coverage_amount_str:
                try:
                    coverage_amount = float(coverage_amount_str)
                except (ValueError, TypeError):
                    coverage_amount = None
                    
            probation_period_str = request.POST.get('probation_period_days', '90')
            try:
                probation_period_days = int(probation_period_str)
            except (ValueError, TypeError):
                probation_period_days = 90  # Default value        

            # Create new employee with all fields
            employee = Employee(
                # Basic Information
                employee_id=employee_id,
                first_name=request.POST.get('first_name'),
                middle_name=request.POST.get('middle_name'),
                last_name=request.POST.get('last_name'),
                email=email,
                
                # Phone numbers
                dial_code=request.POST.get('dial_code', '+91'),
                phone=request.POST.get('phone'),
                alternate_phone=request.POST.get('alternate_phone'),
                residence_number=request.POST.get('residence_number'),
                address = request.POST.get('address'),
                present_address = request.POST.get('present_address'),
                
                # Personal Information
                gender=request.POST.get('gender'),
                date_of_birth=date_of_birth,
                marital_status=request.POST.get('marital_status'),
                marriage_date=marriage_date,
                
                # Family Information
                father_name=request.POST.get('father_name'),
                mother_name=request.POST.get('mother_name'),
                spouse_name=request.POST.get('spouse_name'),
                spouse_gender=request.POST.get('spouse_gender'),
                
                # Medical & Physical Information
                physically_handicapped=physically_handicapped,
                blood_group=request.POST.get('blood_group'),
                nationality=request.POST.get('nationality', 'Indian'),
                
                # Employment Details
                department=department,
                department_id=department_id,
                designation=designation,
                designation_id=designation_id,
                location=location,
                location_id=location_id,
                reporting_manager=reporting_manager_full,
                reporting_manager_id=reporting_manager_id,
                role=request.POST.get('role'),
                date_of_joining=date_of_joining,
                contract_end_date=contract_end_date,
                legal_entity=request.POST.get('legal_entity'),
                worker_type=request.POST.get('worker_type', 'Permanent'),
                status=request.POST.get('status'),
                
                # Profile Picture
                profile_picture=request.FILES.get('profile_picture'),
                
                # Bank & Salary Information
                bank_name=request.POST.get('bank_name'),
                account_number=request.POST.get('account_number'),
                ifsc_code=request.POST.get('ifsc_code'),
                salary_payment_mode=request.POST.get('salary_payment_mode', 'Bank Transfer'),
                name_on_bank_account=request.POST.get('name_on_bank_account'),
                basic_salary=basic_salary,
                
                # PF Details
                pf_establishment_id=request.POST.get('pf_establishment_id'),
                pf_details_available=request.POST.get('pf_details_available', 'No'),
                pf_number=request.POST.get('pf_number'),
                pf_joining_date=pf_joining_date,
                name_on_pf_account=request.POST.get('name_on_pf_account'),
                uan=request.POST.get('uan'),
                
                # ESI Details
                esi_eligible=request.POST.get('esi_eligible', 'No'),
                employer_esi_number=request.POST.get('employer_esi_number'),
                esi_details_available=request.POST.get('esi_details_available', 'No'),
                esi_number=request.POST.get('esi_number'),
                
                # Professional Tax & LWF Details
                pt_establishment_id=request.POST.get('pt_establishment_id'),
                lwf_eligible=request.POST.get('lwf_eligible', 'No'),
                enrollment_number=request.POST.get('enrollment_number'),

                 # Insurance Details
                insurance_type=request.POST.get('insurance_type'),
                policy_name=request.POST.get('policy_name'),
                insurance_company=request.POST.get('insurance_company'),
                policy_number=request.POST.get('policy_number'),
                coverage_amount=coverage_amount,
                nominee_name=request.POST.get('nominee_name'),
                nominee_relationship=request.POST.get('nominee_relationship'),
                policy_start_date=policy_start_date,
                policy_end_date=policy_end_date,
                # probation_period_days=request.POST.get('probation_period_days', 90),
                probation_period_days=probation_period_days,
                created_by=request.session.get('user_name'),
                # Timestamps
                created_at=timezone.now(),
                updated_at=timezone.now()
            )
            
            employee.save()
           
            # Save Educational Certificates
            education_types = request.POST.getlist('education_type[]')
            education_files = request.FILES.getlist('education_files[]')

            for edu_type, edu_file in zip(education_types, education_files):
                if edu_type:  # Save if education type exists
                    EmployeeDocument.objects.create(
                        employee=employee,
                        document_type='educational',
                        document_number=edu_type,
                        file=edu_file if edu_file else None
                    )
           
            # Save PAN Card
            pan_number = request.POST.get('pan_number')
            pan_file = request.FILES.get('pan_file')
            if pan_number:  # Save if pan_number exists
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='pan',
                    document_number=pan_number,
                    file=pan_file  # This can be None
                )
           
            # Save Aadhaar Card
            aadhaar_number = request.POST.get('aadhaar_number')
            aadhaar_file = request.FILES.get('aadhaar_file')
            if aadhaar_number:  # Save if aadhaar_number exists
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='aadhaar',
                    document_number=aadhaar_number,
                    file=aadhaar_file  # This can be None
                )
           
            # Save Bank Passbook
            passbook_file = request.FILES.get('passbook_file')
            if passbook_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='passbook',
                    file=passbook_file
                )
           
            # Save Offer Letter
            offer_letter_file = request.FILES.get('offer_letter_file')
            if offer_letter_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='offer_letter',
                    file=offer_letter_file
                )
           
            # Save Salary Slips (multiple files)
            salary_slip_files = request.FILES.getlist('salary_slip_files')
            for salary_slip_file in salary_slip_files:
                if salary_slip_file:  # Check if file exists
                    EmployeeDocument.objects.create(
                        employee=employee,
                        document_type='salary_slip',
                        file=salary_slip_file
                    )
           
            # Save Bank Statement
            bank_statement_file = request.FILES.get('bank_statement_file')
            if bank_statement_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='bank_statement',
                    file=bank_statement_file
                )
           
            # Save Multiple Experience/Relieving Letters
            experience_companies = request.POST.getlist('experience_company[]')
            experience_letter_files = request.FILES.getlist('experience_letter_files[]')

            print(f"DEBUG: Experience companies: {experience_companies}")
            print(f"DEBUG: Experience files count: {len(experience_letter_files)}")

            # Process each experience letter entry
            for i, (company, exp_file) in enumerate(zip(experience_companies, experience_letter_files)):
                if company or exp_file:  # Save if either company name or file exists
                    # Create a document number that includes company name
                    document_number = f"experience_{company or f'unknown_{i}'}"
                    
                    EmployeeDocument.objects.create(
                        employee=employee,
                        document_type='experience_letter',
                        document_number=document_number,
                        file=exp_file if exp_file else None,
                        # You can add additional metadata if needed
                        # company_name=company  # If you want to store company name separately
                    )
                    print(f"DEBUG: Created experience letter for company: {company}, file: {exp_file.name if exp_file else 'None'}")
            
            # Save Form 16 Document (Simple upload like offer letter)
            form16_file = request.FILES.get('form16_file')
            if form16_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='form16',
                    document_number='Form 16',
                    file=form16_file
                )
                print(f"DEBUG: Created Form 16 document: {form16_file.name}")

            # Save IIR Document (Simple upload like offer letter)
            iir_file = request.FILES.get('iir_file')
            if iir_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='iir',
                    document_number='IIR Investment Declaration',
                    file=iir_file
                )
                print(f"DEBUG: Created IIR document: {iir_file.name}")

            
            messages.success(request, f"Employee {employee.first_name} {employee.middle_name or ''} {employee.last_name} added successfully with all documents!")
            return redirect('employee_page')
           
        except Exception as e:
            messages.error(request, f"Error adding employee: {str(e)}")
            import traceback
            print(traceback.format_exc())  # For debugging
            return render(request, 'hr/add_employee.html', {
                'managers': managers,
                'locations': locations,
                'departments': departments,
                'designations': designations,
                'roles':roles,
            })
   
    # GET request - show empty form with managers data
    context = {
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
        'managers': managers,
        'locations': locations,
        'departments': departments,
        'designations': designations,
        'roles':roles,
        # Add choices for dropdowns
        'gender_choices': Employee.GENDER_CHOICES,
        'marital_status_choices': Employee.MARITAL_STATUS_CHOICES,
        'worker_type_choices': Employee.WORKER_TYPE_CHOICES,
        'salary_payment_mode_choices': Employee.SALARY_PAYMENT_MODE_CHOICES,
        'blood_group_choices': Employee.BLOOD_GROUP_CHOICES,
        'pf_details_choices': Employee.PF_DETAILS_CHOICES,
        'esi_eligible_choices': Employee.ESI_ELIGIBLE_CHOICES,
        'lwf_eligible_choices': Employee.LWF_ELIGIBLE_CHOICES,
        'insurance_type_choices': Employee.INSURANCE_TYPE_CHOICES,
    }
    return render(request, 'hr/add_employee.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'MANAGER', 'SUPER ADMIN' ,'TL', 'BRANCH MANAGER'])
def employee_page(request):
    # Get search query and filters
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    location_filter = request.GET.get('location', '')  # NEW: Location filter
    page_size = int(request.GET.get('page_size', 12))
   
    # Get current user details
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')
    user_name = request.session.get('user_name')
   
    # Start with appropriate employee list based on role
    if user_role in ['ADMIN', 'HR', 'SUPER ADMIN']:
        employees_list = Employee.objects.all().order_by('first_name')
        filter_info = "Showing all employees"
   
    elif user_role == 'MANAGER' or user_role == 'TL':
        try:
            # Get the current manager's employee record
            current_manager = Employee.objects.get(email=user_email)
           
            # Filter by reporting_manager_id OR by reporting_manager name (fallback)
            employees_list = Employee.objects.filter(
                Q(reporting_manager_id=current_manager.id) |
                Q(reporting_manager__icontains=current_manager.first_name)
            ).order_by('first_name')
           
            filter_info = f"Showing employees under {user_name}"
           
        except Employee.DoesNotExist:
            employees_list = Employee.objects.none()
            filter_info = "Manager profile not found"

    elif user_role == 'BRANCH MANAGER':
        try:
            # Get the current branch manager's employee record
            current_branch_manager = Employee.objects.get(email=user_email)
            
            # Get the branch manager's location
            branch_manager_location = current_branch_manager.location
            
            if branch_manager_location:
                # Show all employees with the same location
                employees_list = Employee.objects.filter(
                    location__iexact=branch_manager_location
                ).order_by('first_name')
                
                filter_info = f"Showing employees from {branch_manager_location} location"
            else:
                employees_list = Employee.objects.none()
                filter_info = "Branch manager location not set"
                
        except Employee.DoesNotExist:
            employees_list = Employee.objects.none()
            filter_info = "Branch manager profile not found"
   
    else:
        employees_list = Employee.objects.none()
        filter_info = "No access to employee list"
   
    # Apply search filter (UPDATED: Added location search)
    if search_query:
        employees_list = employees_list.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(department__icontains=search_query) |
            Q(designation__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(employee_id__icontains=search_query) |  # NEW: Search by employee ID
            Q(location__icontains=search_query)       # NEW: Search by location
        )
   
    # Apply department filter
    if department_filter:
        employees_list = employees_list.filter(department__iexact=department_filter)
   
    # Apply status filter
    if status_filter:
        employees_list = employees_list.filter(status__iexact=status_filter)
        
    # NEW: Apply location filter
    if location_filter:
        employees_list = employees_list.filter(location__iexact=location_filter)
   
    # Get total counts for display
    total_employees = employees_list.count()
    active_employees = employees_list.filter(status='active').count()
    
    # NEW: Get unique locations for dropdown
    locations = Employee.objects.values_list('location', flat=True).distinct()
    locations = [loc for loc in locations if loc]  # Remove empty values
    
    # Pagination
    paginator = Paginator(employees_list, page_size)
    page = request.GET.get('page')
   
    try:
        employees = paginator.page(page)
    except PageNotAnInteger:
        employees = paginator.page(1)
    except EmptyPage:
        employees = paginator.page(paginator.num_pages)
   
    context = {
        'employees': employees,
        'today_date': date.today(),
        'search_query': search_query,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'location_filter': location_filter,  # NEW: Location filter
        'locations': locations,              # NEW: Locations for dropdown
        'page_size': page_size,
        'user_name': user_name,
        'user_role': user_role,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'filter_info': filter_info,
    }
   
    return render(request, 'hr/employee.html', context)


@login_required
def employee_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
   
    # Check permission - employees can only view their own profile
    user_role = request.session.get('user_role')
    user_email = request.session.get('user_email')
   
    if user_role == 'EMPLOYEE' and employee.email != user_email:
        messages.error(request, 'You can only view your own profile.')
        return redirect('access_denied')
   
    # Get all active managers for the dropdown (for edit modal)
    managers = Employee.objects.filter(
        role__in=['Manager', 'TL', 'HR', 'Admin', 'Super Admin'],
        status='active'
    ).order_by('first_name', 'last_name')
   
    context = {
        'employee': employee,
        'managers': managers,  # Add this line
        'today_date': timezone.now().date(),
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
    }
   
    return render(request, 'hr/employee_detail.html', context)


@login_required
@role_required(['ADMIN', 'HR', 'MANAGER', 'SUPER ADMIN'])
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
   
    # Get managers for dropdown
    managers = Employee.objects.filter(
        role__in=['Manager', 'TL', 'HR', 'Admin', 'Super Admin'],
        status='active'
    ).order_by('first_name', 'last_name')

    # Get all active locations for the dropdown
    locations = Location.objects.filter(is_active=True).order_by('name')
    
    # Get all active departments for the dropdown
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Get all active designations for the dropdown
    designations = Designation.objects.filter(is_active=True).order_by('title')

    # Get all active designations for the dropdown
    roles = Role.objects.filter(is_active=True).order_by('name')
   
    # Check if manager can edit this employee
    user_role = request.session.get('user_role')
    if user_role == 'MANAGER':
        user_name = request.session.get('user_name')
        if user_name not in employee.reporting_manager:
            messages.error(request, 'You can only edit employees in your team.')
            return redirect('access_denied')
   
    if request.method == 'POST':
        try:
            # Get reporting manager data from form
            reporting_manager_full = request.POST.get('reporting_manager')
            reporting_manager_id = request.POST.get('reporting_manager_id')
           
            # If we have the reporting manager name but no ID, try to find it
            if reporting_manager_full and not reporting_manager_id:
                try:
                    name_part = reporting_manager_full.split(' (')[0]
                    first_name, last_name = name_part.split(' ', 1) if ' ' in name_part else (name_part, '')
                   
                    manager = Employee.objects.filter(
                        first_name=first_name,
                        last_name=last_name,
                        status='active'
                    ).first()
                   
                    if manager:
                        reporting_manager_id = manager.employee_id
                except (IndexError, ValueError):
                    reporting_manager_id = None

            # Handle date fields - FIXED: Use DD-MM-YYYY format
            def parse_date(date_str):
                if date_str:
                    try:
                        # Try parsing as DD-MM-YYYY
                        return datetime.strptime(date_str, "%d-%m-%Y").date()
                    except ValueError:
                        try:
                            # Fallback: try YYYY-MM-DD (in case of browser auto-format)
                            return datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            return None
                return None

            # Handle location
            location_id = request.POST.get('location')
            location_instance = None
            if location_id:
                try:
                    location_instance = Location.objects.get(id=location_id)
                except Location.DoesNotExist:
                    location_instance = None

            # Handle department
            department_id = request.POST.get('department')
            department_instance = None
            if department_id:
                try:
                    department_instance = Department.objects.get(id=department_id)
                except Department.DoesNotExist:
                    department_instance = None

            # Handle designation
            designation_id = request.POST.get('designation')
            designation_instance = None
            if designation_id:
                try:
                    designation_instance = Designation.objects.get(id=designation_id)
                except Designation.DoesNotExist:
                    designation_instance = None

            # Handle basic salary
            basic_salary_str = request.POST.get('basic_salary')
            if basic_salary_str:
                try:
                    employee.basic_salary = float(basic_salary_str)
                except (ValueError, TypeError):
                    employee.basic_salary = 0.00
            else:
                employee.basic_salary = 0.00

            # Handle boolean fields
            physically_handicapped = request.POST.get('physically_handicapped') == 'on'

            # Update employee fields - Personal Information
            employee.employee_id = request.POST.get('employee_id')
            employee.first_name = request.POST.get('first_name')
            employee.middle_name = request.POST.get('middle_name')
            employee.last_name = request.POST.get('last_name')
            employee.email = request.POST.get('email')
            employee.dial_code = request.POST.get('dial_code')
            employee.phone = request.POST.get('phone')
            employee.alternate_phone = request.POST.get('alternate_phone')
            employee.residence_number = request.POST.get('residence_number')
            employee.address = request.POST.get('address')
            employee.present_address = request.POST.get('present_address')
            employee.gender = request.POST.get('gender')
            employee.date_of_birth = parse_date(request.POST.get('date_of_birth'))
            employee.marital_status = request.POST.get('marital_status')
            employee.marriage_date = parse_date(request.POST.get('marriage_date'))
            employee.father_name = request.POST.get('father_name')
            employee.mother_name = request.POST.get('mother_name')
            employee.spouse_name = request.POST.get('spouse_name')
            employee.spouse_gender = request.POST.get('spouse_gender')
            employee.physically_handicapped = physically_handicapped
            employee.blood_group = request.POST.get('blood_group')
            employee.nationality = request.POST.get('nationality')

            # Employment Information
            if department_instance:
                employee.department = department_instance.name
                employee.department_id = department_instance.id
            else:
                employee.department = request.POST.get('department')
                employee.department_id = None

            if designation_instance:
                employee.designation = designation_instance.title
                employee.designation_id = designation_instance.id
            else:
                employee.designation = request.POST.get('designation')
                employee.designation_id = None

            if location_instance:
                employee.location = location_instance.name
                employee.location_id = location_instance.id
            else:
                employee.location = request.POST.get('location')
                employee.location_id = None

            employee.reporting_manager = reporting_manager_full
            employee.reporting_manager_id = reporting_manager_id
            employee.role = request.POST.get('role')
            employee.worker_type = request.POST.get('worker_type')
            employee.date_of_joining = parse_date(request.POST.get('date_of_joining'))
            # employee.probation_period_days=request.POST.get('probation_period_days', 90)
            employee.notice_period_days = int(request.POST.get('notice_period_days', 60))
            # Handle probation period days - convert to integer
            probation_period_str = request.POST.get('probation_period_days', '90')
            try:
                employee.probation_period_days = int(probation_period_str)
            except (ValueError, TypeError):
                employee.probation_period_days = 90  # Default value
            
            
            
            
            employee.contract_end_date = parse_date(request.POST.get('contract_end_date'))
            employee.legal_entity = request.POST.get('legal_entity')
            employee.status = request.POST.get('status')

            # Bank & Salary Information
            employee.bank_name = request.POST.get('bank_name')
            employee.account_number = request.POST.get('account_number')
            employee.ifsc_code = request.POST.get('ifsc_code')
            employee.salary_payment_mode = request.POST.get('salary_payment_mode')
            employee.name_on_bank_account = request.POST.get('name_on_bank_account')

            # PF Details
            employee.pf_establishment_id = request.POST.get('pf_establishment_id')
            employee.pf_details_available = request.POST.get('pf_details_available')
            employee.pf_number = request.POST.get('pf_number')
            employee.pf_joining_date = parse_date(request.POST.get('pf_joining_date'))
            employee.name_on_pf_account = request.POST.get('name_on_pf_account')
            employee.uan = request.POST.get('uan')

            # ESI Details
            employee.esi_eligible = request.POST.get('esi_eligible')
            employee.employer_esi_number = request.POST.get('employer_esi_number')
            employee.esi_details_available = request.POST.get('esi_details_available')
            employee.esi_number = request.POST.get('esi_number')

            # Professional Tax & LWF Details
            employee.pt_establishment_id = request.POST.get('pt_establishment_id')
            employee.lwf_eligible = request.POST.get('lwf_eligible')
            employee.enrollment_number = request.POST.get('enrollment_number')

            # Insurance Details
            employee.insurance_type = request.POST.get('insurance_type')
            employee.policy_name = request.POST.get('policy_name')
            employee.insurance_company = request.POST.get('insurance_company')
            employee.policy_number = request.POST.get('policy_number')
            employee.nominee_name = request.POST.get('nominee_name')
            employee.nominee_relationship = request.POST.get('nominee_relationship')
            employee.policy_start_date = parse_date(request.POST.get('policy_start_date'))
            employee.policy_end_date = parse_date(request.POST.get('policy_end_date'))
            
            # Handle coverage amount
            coverage_amount_str = request.POST.get('coverage_amount')
            if coverage_amount_str:
                try:
                    employee.coverage_amount = float(coverage_amount_str)
                except (ValueError, TypeError):
                    employee.coverage_amount = None
            else:
                employee.coverage_amount = None
           
            # Profile picture - only update if new file is provided
            if request.FILES.get('profile_picture'):
                employee.profile_picture = request.FILES.get('profile_picture')
            employee.updated_by = request.session.get('user_name')
            employee.updated_at = timezone.now()
            employee.save()
           
            # Handle document uploads
            handle_document_uploads(employee, request)
           
            messages.success(request, 'Employee and documents updated successfully!')
            return redirect('employee_detail', employee_id=employee_id)
           
        except Exception as e:
            messages.error(request, f'Error updating employee: {str(e)}')
            import traceback
            print(traceback.format_exc())  # For debugging
   
    # Add choice fields to context for template dropdowns
    context = {
        'employee': employee,
        'managers': managers,
        'locations': locations,
        'departments': departments,
        'designations': designations,
        'roles': roles,
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
        # Add choices for dropdowns
        'gender_choices': Employee.GENDER_CHOICES,
        'marital_status_choices': Employee.MARITAL_STATUS_CHOICES,
        'worker_type_choices': Employee.WORKER_TYPE_CHOICES,
        'salary_payment_mode_choices': Employee.SALARY_PAYMENT_MODE_CHOICES,
        'blood_group_choices': Employee.BLOOD_GROUP_CHOICES,
        'pf_details_choices': Employee.PF_DETAILS_CHOICES,
        'esi_eligible_choices': Employee.ESI_ELIGIBLE_CHOICES,
        'lwf_eligible_choices': Employee.LWF_ELIGIBLE_CHOICES,
        'insurance_type_choices': Employee.INSURANCE_TYPE_CHOICES,
    }  
    return render(request, 'hr/edit_employee.html', context)

# Admin Management
@login_required
@role_required(['ADMIN', 'SUPER ADMIN'])
def admin_list(request):
    admins = Admin.objects.all().order_by('-created_at')
    
    # Calculate statistics
    active_admins_count = admins.filter(status='active').count()
    inactive_admins_count = admins.filter(status='inactive').count()
    roles_count = admins.values('role').distinct().count()
    
    return render(request, 'hr/admin_list.html', {
        'admins': admins,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
        'active_admins_count': active_admins_count,
        'inactive_admins_count': inactive_admins_count,
        'roles_count': roles_count,
    })

@login_required
@role_required(['ADMIN', 'SUPER ADMIN'])
def admin_create(request):
    if request.method == 'POST':
        form = AdminForm(request.POST)
        if form.is_valid():
            try:
                admin = form.save(commit=False)
                # Set empty string for profile_picture since it's required in model
                admin.profile_picture = ""
                
                admin.password_hash = simple_hash('password123')
                admin.created_at = timezone.now()
                admin.updated_at = timezone.now()
                admin.save()
                messages.success(request, 'Admin created successfully!')
                return redirect('admin_list')
            except Exception as e:
                messages.error(request, f'Error creating admin: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AdminForm()
    
    return render(request, 'hr/admin_form.html', {
        'form': form,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
    })

@login_required
@role_required(['ADMIN', 'SUPER ADMIN'])
def admin_update(request, pk):
    admin = get_object_or_404(Admin, pk=pk)
    form = AdminForm(request.POST or None, instance=admin)
    if form.is_valid():
        admin = form.save(commit=False)
        admin.updated_at = timezone.now()
        admin.save()
        messages.success(request, 'Admin updated successfully!')
        return redirect('admin_list')
    return render(request, 'hr/admin_form.html', {
        'form': form,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
    })

@login_required
@role_required(['ADMIN', 'SUPER ADMIN'])
def admin_delete(request, pk):
    admin = get_object_or_404(Admin, pk=pk)
    if request.method == 'POST':
        admin.delete()
        messages.success(request, 'Admin deleted successfully!')
        return redirect('admin_list')
    return render(request, 'hr/admin_confirm_delete.html', {
        'admin': admin,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
    })

def home(request):
    if request.session.get('user_authenticated'):
        user_role = request.session.get('user_role')
        if user_role in ['ADMIN', 'HR', 'SUPER ADMIN','BRANCH MANAGER']:
            return redirect('dashboard')
        else:
            return redirect('employee_dashboard')
    return redirect('login')

@login_required
def update_employee_profile(request):
    """Allow active employees to update their own profile"""
    user_email = request.session.get('user_email')
    user_role = request.session.get('user_role')
    
    # Only allow employees to update their own profile
    # if user_role != 'EMPLOYEE':
    #     messages.error(request, 'Access denied.')
    #     return redirect('access_denied')
    
    try:
        employee = Employee.objects.get(email=user_email, status='active')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found or inactive.')
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        # Update employee profile
        employee.phone = request.POST.get('phone', employee.phone)
        employee.department = request.POST.get('department', employee.department)
        employee.designation = request.POST.get('designation', employee.designation)
        employee.location = request.POST.get('location', employee.location)
        employee.present_address = request.POST.get('present_address', employee.present_address)
        
        # Handle profile picture upload
        if request.FILES.get('profile_picture'):
            employee.profile_picture = request.FILES.get('profile_picture')
            employee.save()
            # Build absolute URL for the profile picture
            profile_picture_url = request.build_absolute_uri(employee.profile_picture.url)
            request.session['profile_picture'] = profile_picture_url
            
            # Debug print
            print(f"DEBUG: Profile picture absolute URL: {profile_picture_url}") 
        employee.updated_by = request.session.get('user_name')
        employee.updated_at = timezone.now()
        employee.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('employee_dashboard')
    
    context = {
        'employee': employee,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': user_role,
    }
    return render(request, 'hr/update_employee_profile.html', context)


def handle_document_uploads(employee, request):
    print("=== DEBUG: Starting document upload process ===")
    
    # Handle Educational Certificates - NEW APPROACH
    # Get all education-related fields from the form
    education_data = {}
    
    # Process all education type fields
    for key, value in request.POST.items():
        if key.startswith('education_type_'):
            education_id = key.replace('education_type_', '')
            education_data[education_id] = {
                'type': value,
                'file': None
            }
            print(f"DEBUG: Found education type: {education_id} -> {value}")
    
    # Process all education file fields
    for key, file_obj in request.FILES.items():
        if key.startswith('education_file_'):
            education_id = key.replace('education_file_', '')
            if education_id in education_data:
                education_data[education_id]['file'] = file_obj
                print(f"DEBUG: Found education file: {education_id} -> {file_obj.name}")
            else:
                print(f"DEBUG: Found orphan education file: {education_id} -> {file_obj.name}")
    
    print(f"DEBUG: Processed education data: {education_data}")
    
    # Get all existing educational documents
    existing_edu_docs = EmployeeDocument.objects.filter(
        employee=employee, 
        document_type='educational'
    )
    existing_docs_dict = {doc.document_number: doc for doc in existing_edu_docs}
    
    print(f"DEBUG: Existing education documents: {list(existing_docs_dict.keys())}")
    
    # Process each education entry
    submitted_types = set()
    
    for education_id, edu_info in education_data.items():
        edu_type = edu_info['type']
        edu_file = edu_info['file']
        
        if not edu_type:  # Skip if no type selected
            continue
            
        submitted_types.add(edu_type)
        print(f"DEBUG: Processing education: id={education_id}, type={edu_type}, file={edu_file.name if edu_file else 'None'}")
        
        # Check if this education type already exists
        existing_doc = existing_docs_dict.get(edu_type)
        
        if existing_doc:
            print(f"DEBUG: Updating existing document for {edu_type}")
            # Update existing document
            if edu_file:
                existing_doc.file = edu_file
                existing_doc.save()
                print(f"DEBUG: UPDATED {edu_type} with file: {edu_file.name}")
            else:
                print(f"DEBUG: No new file for {edu_type}, keeping existing file")
        else:
            print(f"DEBUG: Creating new document for {edu_type}")
            # Create new document
            if edu_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='educational',
                    document_number=edu_type,
                    file=edu_file
                )
                print(f"DEBUG: CREATED {edu_type} with file: {edu_file.name}")
            else:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='educational',
                    document_number=edu_type
                )
                print(f"DEBUG: CREATED {edu_type} without file")
    
    # Handle deletion of education types that are no longer in the form
    for existing_doc in existing_edu_docs:
        if existing_doc.document_number not in submitted_types:
            print(f"DEBUG: DELETING education document for {existing_doc.document_number} (not in form)")
            existing_doc.delete()

    # Handle Experience/Relieving Letters - MULTIPLE ENTRIES
    experience_data = {}
    
    # Process all experience company fields
    for key, value in request.POST.items():
        if key.startswith('experience_company_'):
            experience_id = key.replace('experience_company_', '')
            experience_data[experience_id] = {
                'company': value,
                'file': None
            }
            print(f"DEBUG: Found experience company: {experience_id} -> {value}")
    
    # Process all experience file fields
    for key, file_obj in request.FILES.items():
        if key.startswith('experience_letter_file_'):
            experience_id = key.replace('experience_letter_file_', '')
            if experience_id in experience_data:
                experience_data[experience_id]['file'] = file_obj
                print(f"DEBUG: Found experience file: {experience_id} -> {file_obj.name}")
            else:
                print(f"DEBUG: Found orphan experience file: {experience_id} -> {file_obj.name}")
    
    print(f"DEBUG: Processed experience data: {experience_data}")
    
    # Get all existing experience documents
    existing_exp_docs = EmployeeDocument.objects.filter(
        employee=employee, 
        document_type='experience_letter'
    )
    existing_exp_dict = {doc.document_number: doc for doc in existing_exp_docs}
    
    print(f"DEBUG: Existing experience documents: {list(existing_exp_dict.keys())}")
    
    # Process each experience entry
    submitted_companies = set()
    
    for experience_id, exp_info in experience_data.items():
        company_name = exp_info['company']
        exp_file = exp_info['file']
        
        if not company_name:  # Skip if no company name
            continue
            
        submitted_companies.add(company_name)
        print(f"DEBUG: Processing experience: id={experience_id}, company={company_name}, file={exp_file.name if exp_file else 'None'}")
        
        # Check if this company already exists
        existing_doc = existing_exp_dict.get(company_name)
        
        if existing_doc:
            print(f"DEBUG: Updating existing experience document for {company_name}")
            # Update existing document
            if exp_file:
                existing_doc.file = exp_file
                existing_doc.save()
                print(f"DEBUG: UPDATED {company_name} with file: {exp_file.name}")
            else:
                print(f"DEBUG: No new file for {company_name}, keeping existing file")
        else:
            print(f"DEBUG: Creating new experience document for {company_name}")
            # Create new document
            if exp_file:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='experience_letter',
                    document_number=company_name,
                    file=exp_file
                )
                print(f"DEBUG: CREATED {company_name} with file: {exp_file.name}")
            else:
                EmployeeDocument.objects.create(
                    employee=employee,
                    document_type='experience_letter',
                    document_number=company_name
                )
                print(f"DEBUG: CREATED {company_name} without file")
    
    # Handle deletion of experience letters that are no longer in the form
    for existing_doc in existing_exp_docs:
        if existing_doc.document_number not in submitted_companies:
            print(f"DEBUG: DELETING experience document for {existing_doc.document_number} (not in form)")
            existing_doc.delete()
    
    # Handle PAN Card (keep existing code)
    pan_number = request.POST.get('pan_number')
    pan_file = request.FILES.get('pan_file')
    
    print(f"DEBUG: PAN Number: {pan_number}, PAN File: {pan_file.name if pan_file else 'None'}")
    
    existing_pan = EmployeeDocument.objects.filter(employee=employee, document_type='pan').first()
    
    if pan_file:
        if existing_pan:
            existing_pan.document_number = pan_number
            existing_pan.file = pan_file
            existing_pan.save()
            print(f"DEBUG: Updated PAN with new file: {pan_file.name}")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='pan',
                document_number=pan_number,
                file=pan_file
            )
            print(f"DEBUG: Created new PAN with file: {pan_file.name}")
    elif pan_number:
        if existing_pan:
            existing_pan.document_number = pan_number
            existing_pan.save()
            print(f"DEBUG: Updated PAN number only")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='pan',
                document_number=pan_number
            )
            print(f"DEBUG: Created new PAN without file")
    
    # Handle Aadhaar Card (keep existing code)
    aadhaar_number = request.POST.get('aadhaar_number')
    aadhaar_file = request.FILES.get('aadhaar_file')
    
    print(f"DEBUG: Aadhaar Number: {aadhaar_number}, Aadhaar File: {aadhaar_file.name if aadhaar_file else 'None'}")
    
    existing_aadhaar = EmployeeDocument.objects.filter(employee=employee, document_type='aadhaar').first()
    
    if aadhaar_file:
        if existing_aadhaar:
            existing_aadhaar.document_number = aadhaar_number
            existing_aadhaar.file = aadhaar_file
            existing_aadhaar.save()
            print(f"DEBUG: Updated Aadhaar with new file: {aadhaar_file.name}")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='aadhaar',
                document_number=aadhaar_number,
                file=aadhaar_file
            )
            print(f"DEBUG: Created new Aadhaar with file: {aadhaar_file.name}")
    elif aadhaar_number:
        if existing_aadhaar:
            existing_aadhaar.document_number = aadhaar_number
            existing_aadhaar.save()
            print(f"DEBUG: Updated Aadhaar number only")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='aadhaar',
                document_number=aadhaar_number
            )
            print(f"DEBUG: Created new Aadhaar without file")

    # Handle Form 16 Document - NEW
    form16_file = request.FILES.get('form16_file')
    if form16_file:
        existing_form16 = EmployeeDocument.objects.filter(employee=employee, document_type='form16').first()
        if existing_form16:
            existing_form16.file = form16_file
            existing_form16.save()
            print(f"DEBUG: Updated Form 16 with new file: {form16_file.name}")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='form16',
                document_number='Form 16',
                file=form16_file
            )
            print(f"DEBUG: Created new Form 16 with file: {form16_file.name}")
    
    # Handle IIR Document - NEW
    iir_file = request.FILES.get('iir_file')
    if iir_file:
        existing_iir = EmployeeDocument.objects.filter(employee=employee, document_type='iir').first()
        if existing_iir:
            existing_iir.file = iir_file
            existing_iir.save()
            print(f"DEBUG: Updated IIR with new file: {iir_file.name}")
        else:
            EmployeeDocument.objects.create(
                employee=employee,
                document_type='iir',
                document_number='IIR Investment Declaration',
                file=iir_file
            )
            print(f"DEBUG: Created new IIR with file: {iir_file.name}")

    
    print("=== DEBUG: Document upload process completed ===")

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def delete_document(request, document_id):
    print(f"Delete document called with ID: {document_id}")  # Debug print
    print(f"Request method: {request.method}")  # Debug print
    
    if request.method == "POST":
        try:
            document = get_object_or_404(EmployeeDocument, id=document_id)
            employee_id = document.employee.id
            print(f"Document found: {document}, Employee ID: {employee_id}")  # Debug print
            document.delete()
            messages.success(request, "Document deleted successfully!")
            return redirect('employee_detail', employee_id=employee_id)
        except Exception as e:
            print(f"Error: {str(e)}")  # Debug print
            messages.error(request, f"Error deleting document: {str(e)}")
            return redirect('employee_page')
   
    # If not POST method, redirect to employee page
    messages.error(request, "Invalid request method.")
    return redirect('employee_page')
# All employee
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def all_employee(request):
    employees = Employee.objects.all()
    if request.GET.get('download') == 'excel':
        return download_employees_excel(employees)
    
    return render(request, 'hr/all_employee.html', {
        'employees': employees,
        'today_date': date.today(),
    })

def download_employees_excel(employees):
    """Generate Excel file with all employee data"""
    
    # Create a workbook and add worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Employees"
    
    # Define headers
    headers = [
        'Employee ID', 
        'First Name', 
        'Last Name', 
        'Email', 
        'Phone', 
        'Department', 
        'Location',
        'Status'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Write employee data
    for row_num, employee in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=employee.employee_id)
        ws.cell(row=row_num, column=2, value=employee.first_name)
        ws.cell(row=row_num, column=3, value=employee.last_name)
        ws.cell(row=row_num, column=4, value=employee.email)
        ws.cell(row=row_num, column=5, value=employee.phone)
        ws.cell(row=row_num, column=6, value=employee.department)
        ws.cell(row=row_num, column=7, value=employee.location)
        ws.cell(row=row_num, column=8, value=employee.status)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        column = ws[column_letter]
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="all_employees.xlsx"'
    
    wb.save(response)
    return response

# Active employee
@login_required
def active_employee(request):
    employees = Employee.objects.filter(status__iexact='active')  

    return render(request, 'hr/active_employee.html', {
        'employees': employees,
        'today_date': date.today(),
    })
# ---------------------------------------------------------
# Location Management Views
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def location_list(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    branches = Location.objects.all().order_by('name')
    
    # Apply search filter
    if search_query:
        branches = branches.filter(
            Q(name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(state__icontains=search_query) |
            Q(country__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        if status_filter == 'active':
            branches = branches.filter(is_active=True)
        elif status_filter == 'inactive':
            branches = branches.filter(is_active=False)
    
    context = {
        'branches': branches,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'hr/master_data/location_list.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def location_create(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            location = form.save()
            messages.success(request, f'Location "{location.name}" created successfully!')
            return redirect('location_list')
        else:
            # Return to the same page with form errors
            branches = Location.objects.all().order_by('name')
            context = {
                'branches': branches,
                'form': form,
                'show_add_modal': True  # Flag to show add modal with errors
            }
            return render(request, 'hr/master_data/location_list.html', context)
    
    return redirect('location_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def location_edit(request, pk):
    location = get_object_or_404(Location, pk=pk)
    
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            location = form.save()
            messages.success(request, f'Location "{location.name}" updated successfully!')
            return redirect('location_list')
        else:
            # Return to the same page with form errors
            branches = Location.objects.all().order_by('name')
            context = {
                'branches': branches,
                'form': form,
                'editing_location_id': pk,  # Flag to show edit modal with errors
                'editing_location': location
            }
            return render(request, 'hr/master_data/location_list.html', context)
    
    return redirect('location_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def location_delete(request, pk):
    location = get_object_or_404(Location, pk=pk)
    
    if request.method == 'POST':
        location_name = location.name
        location.delete()
        messages.success(request, f'Location "{location_name}" deleted successfully!')
        return redirect('location_list')
    
    return redirect('location_list')

# Department Management Views
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def department_list(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    departments = Department.objects.all()

    if search_query:
        departments = departments.filter(name__icontains=search_query)
    if status_filter:
        departments = departments.filter(is_active=(status_filter == 'active'))

    context = {
        'departments': departments,
        'search_query': search_query,
        'status_filter': status_filter,
        'form': DepartmentForm(),   # blank form by default
    }
    return render(request, 'hr/master_data/department_list.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def department_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department "{department.name}" created successfully!')
            return redirect('department_list')
        else:
            # Return to the same page with form errors
            departments = Department.objects.all().order_by('name')
            context = {
                'departments': departments,
                'form': form,
                'show_add_modal': True  # Flag to show add modal with errors
            }
            return render(request, 'hr/master_data/department_list.html', context)
    
    return redirect('department_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Department "{department.name}" updated successfully!')
            return redirect('department_list')
        else:
            # Return to the same page with form errors
            departments = Department.objects.all().order_by('name')
            context = {
                'departments': departments,
                'form': form,
                'editing_department_id': pk,  # Flag to show edit modal with errors
                'editing_department': department
            }
            return render(request, 'hr/master_data/department_list.html', context)
    
    return redirect('department_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        department_name = department.name
        department.delete()
        messages.success(request, f'Department "{department_name}" deleted successfully!')
        return redirect('department_list')
    
    return redirect('department_list')

# Designation Management Views
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def designation_list(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    designations = Designation.objects.select_related('department').all().order_by('department__name', 'level')
    departments = Department.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        designations = designations.filter(
            Q(title__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(department_name_icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        if status_filter == 'active':
            designations = designations.filter(is_active=True)
        elif status_filter == 'inactive':
            designations = designations.filter(is_active=False)
    
    context = {
        'designations': designations,
        'departments': departments,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'hr/master_data/designation_list.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def designation_create(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            designation = form.save()
            messages.success(request, f'Designation "{designation.title}" created successfully!')
            return redirect('designation_list')
        else:
            # Return to the same page with form errors
            designations = Designation.objects.select_related('department').all().order_by('department__name', 'level')
            departments = Department.objects.filter(is_active=True)
            context = {
                'designations': designations,
                'departments': departments,
                'form': form,
                'show_add_modal': True  # Flag to show add modal with errors
            }
            return render(request, 'hr/master_data/designation_list.html', context)
    
    return redirect('designation_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def designation_edit(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    
    if request.method == 'POST':
        form = DesignationForm(request.POST, instance=designation)
        if form.is_valid():
            designation = form.save()
            messages.success(request, f'Designation "{designation.title}" updated successfully!')
            return redirect('designation_list')
        else:
            # Return to the same page with form errors
            designations = Designation.objects.select_related('department').all().order_by('department__name', 'level')
            departments = Department.objects.filter(is_active=True)
            context = {
                'designations': designations,
                'departments': departments,
                'form': form,
                'editing_designation_id': pk,  # Flag to show edit modal with errors
                'editing_designation': designation
            }
            return render(request, 'hr/master_data/designation_list.html', context)
    
    return redirect('designation_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def designation_delete(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    
    if request.method == 'POST':
        designation_title = designation.title
        designation.delete()
        messages.success(request, f'Designation "{designation_title}" deleted successfully!')
        return redirect('designation_list')
    
    return redirect('designation_list')


@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def role_list(request):
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    # Start with all roles
    roles = Role.objects.all().order_by('-created_at')

    # Apply search filter
    if search_query:
        roles = roles.filter(
            Q(name__icontains=search_query)
        )

    # Apply status filter
    if status_filter == 'active':
        roles = roles.filter(is_active=True)
    elif status_filter == 'inactive':
        roles = roles.filter(is_active=False)

    context = {
        'roles': roles,
        'search_query': search_query,
        'status_filter': status_filter,
        'title': 'Role Management'
    }
    return render(request, 'hr/master_data/role_list.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def role_create(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            role = form.save()
            messages.success(request, f'Role "{role.name}" created successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('role_list')
        else:
            # Return form errors for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {field: [error for error in error_list] for field, error_list in form.errors.items()}
                return JsonResponse({'success': False, 'errors': errors})
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RoleForm()

    return redirect('role_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def role_edit(request, pk):
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            role = form.save()
            messages.success(request, f'Role "{role.name}" updated successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            return redirect('role_list')
        else:
            # Return form errors for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {field: [error for error in error_list] for field, error_list in form.errors.items()}
                return JsonResponse({'success': False, 'errors': errors})
            messages.error(request, 'Please correct the errors below.')
    
    return redirect('role_list')

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def role_delete(request, pk):
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':         
        role_name = role.name
        role.delete()
        messages.success(request, f'Role "{role_name}" deleted successfully!')
        return redirect('role_list')

    return redirect('role_list')
# ----------------------------------------------------------
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def probation_settings(request):
    # Get current active configuration
    config = ProbationConfiguration.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        days = request.POST.get('probation_period_days')
        if days and days.isdigit():
            try:
                # Deactivate current active configuration
                ProbationConfiguration.objects.update(is_active=False)
                
                # Create or update new active configuration
                if config:
                    config.probation_period_days = int(days)
                    config.is_active = True
                    config.save()
                else:
                    ProbationConfiguration.objects.create(
                        probation_period_days=int(days),
                        is_active=True
                    )
                
                messages.success(request, f'Probation period updated to {days} days')
                return redirect('probation_settings')
                
            except Exception as e:
                messages.error(request, f'Error updating probation period: {str(e)}')
        else:
            messages.error(request, 'Please enter a valid number of days')
    
    context = {
        'config': config,
        'today_date': date.today(),
        'user_name': request.session.get('user_name'),
        'user_role': request.session.get('user_role'),
    }
    return render(request, 'hr/probation_settings.html', context)


@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def warning_list(request):

    warnings = EmployeeWarning.objects.all().order_by('-created_at')
    message_categories = MessageCategory.objects.filter(is_active=True)
    
    warning_count = warnings.filter(message_category="Warning").count()

    appreciation_count = warnings.filter(
        message_category__in=["Appreciation", "Appreciations"]
    ).count()

    notice_count = warnings.exclude(
        message_category__in=["Warning", "Appreciation", "Appreciations"]
    ).count()

    total_count = warnings.count()
    
    if request.method == "POST":
        form = EmployeeWarningForm(request.POST)
        if form.is_valid():
            warning = form.save(commit=False)

            # Category
            cat_id = request.POST.get("message_category")
            cat_obj = MessageCategory.objects.get(id=cat_id)
            warning.message_category = cat_obj.name

            # Subtype
            subtype_id = request.POST.get("sub_type")
            subtype_obj = MessageSubType.objects.get(id=subtype_id)
            warning.sub_type = subtype_obj.name   # store name

            warning.issued_by = request.session.get("user_name")
            warning.save()

            messages.success(request, "Notice added successfully!")
            return redirect("warning_list")

        else:
            print("FORM ERRORS:", form.errors)

    else:
        form = EmployeeWarningForm()

    return render(request, "hr/warning_list.html", {
        "warnings": warnings,
        "form": form,
        "message_categories": message_categories,
        "warning_count": warning_count,
        "appreciation_count": appreciation_count,
        "notice_count": notice_count,
        "total_count": total_count,
    })

def add_warning(request):
    if request.method == 'POST':
        form = EmployeeWarningForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Warning added successfully!")
            return redirect('warning_list')
    else:
        form = EmployeeWarningForm()
    return render(request, 'hr/add_warning.html', {'form': form}) 
    
@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def delete_warning(request, warning_id):
    warning = get_object_or_404(EmployeeWarning, id=warning_id)
    warning.delete()
    messages.success(request, "Warning deleted successfully!")
    return redirect('warning_list')

def search_managers(request):
    """
    Search managers by ID or name — return only employee_id and name
    """
    if request.method == 'POST':
        search_term = request.POST.get('search_term', '').strip()

        if not search_term:
            return JsonResponse({'managers': []})

        # Filter active employees (optionally limit to managers)
        employees = Employee.objects.filter(
            role__in=['Manager', 'HR', 'Admin', 'Super Admin', 'TL'],
            status='active'
        ).filter(
            Q(employee_id__icontains=search_term) |
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term)
        )

        # Return only required fields
        managers_data = [{
            'employee_id': emp.employee_id,
            'name': f"{emp.first_name} {emp.last_name}"
        } for emp in employees[:10]]

        return JsonResponse({'managers': managers_data})

    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
def search_employees(request):
    """
    Search employees by ID or name
    """
    if request.method == 'POST':
        search_term = request.POST.get('search_term', '').strip()
        
        if not search_term:
            return JsonResponse({'employees': []})
        
        # Search by employee_id or first_name or last_name
        employees = Employee.objects.filter(
            status='active'
        ).filter(
            employee_id__icontains=search_term
        ) | Employee.objects.filter(
            status='active'
        ).filter(
            first_name__icontains=search_term
        ) | Employee.objects.filter(
            status='active'
        ).filter(
            last_name__icontains=search_term
        )
        
        employees_data = []
        for emp in employees[:10]:
            employees_data.append({
                'id': emp.id,
                'employee_id': emp.employee_id,
                'name': f"{emp.first_name} {emp.last_name}",
                'department': emp.department,
                'designation': emp.designation
            })
        
        return JsonResponse({'employees': employees_data})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
@role_required(['ADMIN', 'SUPER ADMIN'])
def permission_center(request):
    """
    Render permission center page
    """
    return render(request, 'permission_center.html')




def get_roles(request):
    """
    Get all active roles
    """
    roles = Role.objects.filter(is_active=True)
    roles_data = [{'id': role.id, 'name': role.name} for role in roles]
    return JsonResponse({'roles': roles_data})
def get_all_menus(request):
    """
    Get all menus (both with and without submenus) for dropdown
    """
    try:
        # Get all active menus
        menus = YsMenuMaster.objects.filter(status=True).order_by('seq')
        
        menu_structure = []
        for menu in menus:
            # Get active menu links for this menu
            menu_links = YsMenuLinkMaster.objects.filter(
                menu_id=menu.menu_id, 
                status=1
            ).order_by('seq')
            
            menu_data = {
                'menu_id': menu.menu_id,
                'menu_name': menu.menu_name,
                'menu_url': menu.menu_url,  # Add menu URL for standalone menus
                'has_submenus': menu_links.exists(),
                'menu_links': []
            }
            
            # Add menu links if they exist
            for link in menu_links:
                menu_data['menu_links'].append({
                    'menu_link_id': link.menu_link_id,
                    'menu_link_name': link.menu_link_name,
                    'menu_link_url': link.menu_link_url
                })
            
            menu_structure.append(menu_data)
        
        return JsonResponse({'menus': menu_structure})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_assigned_permissions(request):
    """
    Get already assigned permissions for a role (both menus and submenus)
    """
    role_id = request.GET.get('role_id')
    
    if not role_id:
        return JsonResponse({'assigned_menus': []})
    
    try:
        # Get assigned menu links for this role
        assigned_permissions = YsMenuRoleMaster.objects.filter(
            userRoleId=role_id,
            status=True
        )
        
        assigned_items = []
        for perm in assigned_permissions:
            # Check if this is a standalone menu (menu_link_id equals menu_id)
            if perm.menu_link_id == perm.menu_id:
                # This is likely a standalone menu
                assigned_items.append(f"menu_{perm.menu_id}")
            else:
                # This is a submenu
                assigned_items.append(str(perm.menu_link_id))
        
        print(f"Assigned items for role {role_id}: {assigned_items}")  # Debug log
        
        return JsonResponse({
            'assigned_menus': assigned_items
        })
    
    except Exception as e:
        print(f"Error in get_assigned_permissions: {str(e)}")  # Debug log
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def assign_permissions(request):
    """
    Assign permissions to role (both menus and submenus)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            role_id = data.get('role_id')
            menu_links = data.get('menu_links', [])
            
            print(f"Received data: role_id={role_id}, menu_links={menu_links}")  # Debug log
            
            # Remove existing permissions for this role
            YsMenuRoleMaster.objects.filter(userRoleId=role_id).update(status=False)
            
            # Add new permissions
            for item_id in menu_links:
                print(f"Processing item: {item_id}, type: {type(item_id)}")  # Debug log
                
                if isinstance(item_id, str) and item_id.startswith('menu_'):
                    # This is a standalone menu (not a submenu)
                    menu_id = int(item_id.replace('menu_', ''))
                    print(f"Standalone menu - menu_id: {menu_id}")  # Debug log
                    
                    # For standalone menus, use filter().first() instead of get() to handle duplicates
                    existing_permission = YsMenuRoleMaster.objects.filter(
                        userRoleId=role_id,
                        menu_link_id=menu_id,
                        menu_id=menu_id
                    ).first()
                    
                    if existing_permission:
                        # Update existing permission
                        existing_permission.status = True
                        existing_permission.save()
                    else:
                        # Create new permission
                        YsMenuRoleMaster.objects.create(
                            userRoleId=role_id,
                            menu_link_id=menu_id,
                            menu_id=menu_id,
                            status=True
                        )
                        
                else:
                    # This is a regular submenu
                    try:
                        menu_link_id = int(item_id)
                        print(f"Submenu - menu_link_id: {menu_link_id}")  # Debug log
                        
                        # Get menu_id from the menu link
                        menu_link = YsMenuLinkMaster.objects.get(menu_link_id=menu_link_id)
                        
                        # Use filter().first() instead of get() to handle duplicates
                        existing_permission = YsMenuRoleMaster.objects.filter(
                            userRoleId=role_id,
                            menu_link_id=menu_link_id
                        ).first()
                        
                        if existing_permission:
                            # Update existing permission
                            existing_permission.menu_id = menu_link.menu_id
                            existing_permission.status = True
                            existing_permission.save()
                        else:
                            # Create new permission
                            YsMenuRoleMaster.objects.create(
                                userRoleId=role_id,
                                menu_link_id=menu_link_id,
                                menu_id=menu_link.menu_id,
                                status=True
                            )
                            
                    except (ValueError, TypeError) as e:
                        print(f"Error processing submenu item {item_id}: {e}")
                        continue
                    except YsMenuLinkMaster.DoesNotExist:
                        print(f"Menu link {item_id} does not exist")
                        continue
            
            return JsonResponse({'success': True, 'message': 'Permissions assigned successfully!'})
        
        except Exception as e:
            print(f"Error in assign_permissions: {str(e)}")  # Debug log
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

# domain permission
@login_required
@role_required(['SUPER ADMIN'])
def domain_management(request):
    """Main domain management page"""
    domains = AllowedDomain.objects.all().order_by('domain')
    
    # Initialize forms
    add_form = AllowedDomainForm(request.POST or None, prefix='add')
    edit_form = AllowedDomainForm(prefix='edit')
    
    if request.method == 'POST' and 'add_domain' in request.POST:
        if add_form.is_valid():
            add_form.save()
            messages.success(request, f'Domain {add_form.cleaned_data["domain"]} added successfully!')
            return redirect('domain_management')
        else:
            messages.error(request, 'Please correct the errors below.')
    
    context = {
        'domains': domains,
        'add_form': add_form,
        'edit_form': edit_form,
        'active_page': 'domain_management'
    }
    return render(request, 'domain_management.html', context)



@login_required
@role_required(['SUPER ADMIN'])
@require_http_methods(["POST"])
def add_domain(request):
    """Add new domain"""
    form = AllowedDomainForm(request.POST)
    if form.is_valid():
        domain = form.save()
        messages.success(request, f'Domain {domain.domain} added successfully!')
    else:
        for error in form.errors.values():
            messages.error(request, error)
    
    return redirect('domain_management')

@login_required
@role_required(['SUPER ADMIN'])
@require_http_methods(["POST"])
def update_domain(request, domain_id):
    """Update existing domain"""
    domain = get_object_or_404(AllowedDomain, id=domain_id)
    form = AllowedDomainForm(request.POST, instance=domain)
    
    if form.is_valid():
        form.save()
        messages.success(request, f'Domain {domain.domain} updated successfully!')
    else:
        for error in form.errors.values():
            messages.error(request, error)
    
    return redirect('domain_management')

@login_required
@role_required(['SUPER ADMIN'])
def toggle_domain_status(request, domain_id):
    """Toggle domain active/inactive status"""
    domain = get_object_or_404(AllowedDomain, id=domain_id)
    domain.is_active = not domain.is_active
    domain.save()
    
    status = "activated" if domain.is_active else "deactivated"
    messages.success(request, f'Domain {domain.domain} {status} successfully!')
    
    return redirect('domain_management')

@login_required
@role_required(['SUPER ADMIN'])
def delete_domain(request, domain_id):
    """Delete domain"""
    domain = get_object_or_404(AllowedDomain, id=domain_id)
    domain_name = domain.domain
    domain.delete()
    
    messages.success(request, f'Domain {domain_name} deleted successfully!')
    return redirect('domain_management')

@login_required
@role_required(['SUPER ADMIN'])
def get_domain_details(request, domain_id):
    """Get domain details for editing (AJAX)"""
    domain = get_object_or_404(AllowedDomain, id=domain_id)
    return render(request, 'domain_edit_form.html', {'domain': domain})




# wishes


@csrf_exempt
@login_required
def send_celebration_wish(request):
    """Send celebration wish to a celebrant"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            celebrant_id = data.get('celebrant_id')
            message = data.get('message')
            wish_type = data.get('wish_type')
            
            # Get wisher (current user)
            user_email = request.session.get('user_email')
            try:
                wisher = Employee.objects.get(email=user_email, status='active')
                celebrant = Employee.objects.get(id=celebrant_id, status='active')
                
                # Create wish
                wish = CelebrationWish.objects.create(
                    celebrant=celebrant,
                    wisher=wisher,
                    message=message,
                    wish_type=wish_type
                )
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Wish sent successfully!',
                    'wish_id': wish.id
                })
                
            except Employee.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'error': 'Employee not found'
                }, status=404)
                
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def get_celebration_wishes(request, celebrant_id):
    # Get wish_type from query parameters
    wish_type = request.GET.get('wish_type')
    
    # Filter wishes by celebrant AND type if provided
    wishes_query = CelebrationWish.objects.filter(celebrant_id=celebrant_id)
    
    if wish_type:
        wishes_query = wishes_query.filter(wish_type=wish_type)
    
    # Order by creation date (newest first)
    wishes = wishes_query.order_by('-created_at')
    
    # Serialize wishes data
    wishes_data = []
    for wish in wishes:
        # Convert UTC time to local timezone
        local_time = localtime(wish.created_at)
        
        wishes_data.append({
           'wisher_name': wish.wisher.full_name,
           'wisher_designation': wish.wisher.designation,
           'message': wish.message,
           'time_ago': local_time.strftime('%I:%M %p'),
           'wish_type': wish.get_wish_type_display()      
        })
    
    return JsonResponse({
        'wishes': wishes_data,
        'total_count': len(wishes_data)
    })
    
    
# Add this new AJAX view function
@login_required
def get_designations_by_department(request):
    """AJAX view to get designations by department"""
    department_id = request.GET.get('department_id')
    
    if department_id:
        designations = Designation.objects.filter(
            department_id=department_id, 
            is_active=True
        ).order_by('title')
    else:
        designations = Designation.objects.filter(is_active=True).order_by('title')
    
    designations_data = []
    for designation in designations:
        designations_data.append({
            'id': designation.id,
            'title': designation.title,
            'level': designation.level or '',
            'display_text': f"{designation.title} {f'(Level {designation.level})' if designation.level else ''}"
        })
    
    return JsonResponse({'designations': designations_data})

@login_required
def employee_search_ajax(request):
    search = request.GET.get("search", "")

    employees = Employee.objects.filter(
        Q(first_name__icontains=search) |
        Q(last_name__icontains=search) |
        Q(employee_id__icontains=search) |
        Q(email__icontains=search) |
        Q(phone__icontains=search) |
        Q(department__icontains=search) |
        Q(location__icontains=search)
    ).order_by("first_name")

    data = []

    for emp in employees:
        data.append({
            "id": emp.id,
            "first_name": emp.first_name,
            "middle_name": emp.middle_name,
            "last_name": emp.last_name,
            "employee_id": emp.employee_id,
            "email": emp.email,
            "phone": emp.phone,
            "department": emp.department,
            "designation": emp.designation,
            "location": emp.location,
            "status": emp.status,
            "profile_picture": emp.profile_picture.url if emp.profile_picture else "/static/default.png",
        })

    return JsonResponse({"results": data})


def warning_master_list(request):
    search_query = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    categories = MessageCategory.objects.all()

    if search_query:
        categories = categories.filter(name__icontains=search_query)

    if status_filter == "active":
        categories = categories.filter(is_active=True)
    elif status_filter == "inactive":
        categories = categories.filter(is_active=False)

    categories = categories.order_by('-id')

    return render(request, "hr/master_data/warning.html", {
        "categories": categories,
        "search_query": search_query,
        "status_filter": status_filter,
    })



def warning_master_delete(request, pk):
    category = get_object_or_404(MessageCategory, pk=pk)
    category.delete()
    messages.success(request, "Message Type deleted successfully!")
    return redirect("warning_master_list")



def load_subtypes(request):
    category_id = request.GET.get("category_id")
    subtypes = MessageSubType.objects.filter(category_id=category_id, is_active=True)
    return JsonResponse(list(subtypes.values("id", "name")), safe=False)


def message_category_list(request):
    categories = MessageCategory.objects.all()
    return render(request, "hr/master_data/message_category_list.html", {"categories": categories})


def message_category_create(request):
    if request.method == "POST":
        category_name = request.POST.get("name")
        subtype_name = request.POST.get("subtype_name")

        # 1️⃣ Create or fetch category
        category, created = MessageCategory.objects.get_or_create(
            name=category_name,
            defaults={"is_active": True}
        )

        # 2️⃣ Create subtype
        if subtype_name:
            MessageSubType.objects.create(
                category=category,
                name=subtype_name,
                is_active=True
            )

        messages.success(request, "Message Type & Subtype Saved Successfully!")
        return redirect("warning_master_list")


def message_subtype_list(request, category_id):
    category = get_object_or_404(MessageCategory, id=category_id)
    subtypes = MessageSubType.objects.filter(category=category)

    return render(request, "hr/master_data/message_subtype_list.html", {
        "category": category,
        "subtypes": subtypes
    })


def message_subtype_create(request):
    if request.method == "POST":
        category_id = request.POST.get("category_id")
        name = request.POST.get("name")

        MessageSubType.objects.create(category_id=category_id, name=name)
        return redirect("message_subtype_list", category_id=category_id)


# ===========================
# MESSAGE CATEGORY – EDIT
# ===========================
def message_category_edit(request, pk):
    category = get_object_or_404(MessageCategory, id=pk)

    if request.method == "POST":
        category.name = request.POST.get("name")
        category.is_active = request.POST.get("is_active") == "on"

        # 🔥 NEW – update subtypes
        subtype_string = request.POST.get("subtypes", "")
        subtype_list = [x.strip() for x in subtype_string.split(",") if x.strip()]

        # delete previous subtypes
        category.subtypes.all().delete()

        # create new subtypes
        for st in subtype_list:
            MessageSubType.objects.create(category=category, name=st)

        category.save()
        messages.success(request, "Message Category updated successfully!")
        return redirect("warning_master_list")

    return redirect("warning_master_list")




# ===========================
# MESSAGE CATEGORY – DELETE
# ===========================
def message_category_delete(request, pk):
    category = get_object_or_404(MessageCategory, pk=pk)
    category.delete()
    messages.success(request, "Message Type deleted successfully!")
    return redirect("warning_master_list")


# ===========================
# MESSAGE SUBTYPE – EDIT
# ===========================
def message_subtype_edit(request, pk):
    subtype = get_object_or_404(MessageSubType, pk=pk)

    if request.method == "POST":
        subtype.name = request.POST.get("name")
        subtype.is_active = request.POST.get("is_active") == "on"
        subtype.save()
        messages.success(request, "Subtype updated successfully!")
        return redirect("warning_master_list")

    return redirect("warning_master_list")


# ===========================
# MESSAGE SUBTYPE – DELETE
# ===========================
def message_subtype_delete(request, pk):
    subtype = get_object_or_404(MessageSubType, pk=pk)
    subtype.delete()
    messages.success(request, "Subtype deleted successfully!")
    return redirect("warning_master_list")